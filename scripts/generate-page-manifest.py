import json
import re
import sys
from pathlib import Path

from openpyxl import load_workbook


REQUIRED_COLUMNS = [
    "ID",
    "Publishing_Wave",
    "URL",
    "Path",
    "Category",
    "Tier",
    "Priority",
    "Board",
    "Class",
    "Subject",
    "Location_Type",
    "Location_Name",
    "Primary_Keyword",
    "Search_Intent",
    "Word_Count_Target",
    "Indexation_Advice",
    "Parent_Hub",
    "Link_1",
    "Link_2",
    "Link_3",
    "Link_4",
    "Link_5",
    "Suggested_Schema",
    "Content_Angle",
    "Uniqueness_Block",
    "Source_URL",
]


def normalize_path(value: str | None) -> str | None:
    if value is None:
        return None

    text = str(value).strip()
    if not text:
        return None

    text = re.sub(r"^https?://[^/]+", "", text, flags=re.IGNORECASE)

    if not text.startswith("/"):
        text = f"/{text}"

    text = text.replace("\\", "/")
    text = re.sub(r"/+", "/", text)

    if len(text) > 1 and text.endswith("/"):
        text = text[:-1]

    return text.lower()


def route_family(path: str) -> str:
    segments = [segment for segment in path.split("/") if segment]

    if not segments:
        return "core"

    root = segments[0]

    if root == "boards":
        if len(segments) == 1:
            return "boards_hub"
        if len(segments) == 2:
            return "boards_board"
        if len(segments) == 3:
            return "boards_class"
        if len(segments) == 4:
            return "boards_subject_service"
        return "boards_other"

    if root == "classes":
        if len(segments) == 1:
            return "classes_hub"
        if len(segments) == 2:
            return "classes_overview"
        if len(segments) == 3:
            return "classes_service"
        return "classes_other"

    if root == "resources":
        if len(segments) == 1:
            return "resources_hub"
        if len(segments) == 2:
            return "resources_category"
        if len(segments) == 3:
            return "resources_article"
        return "resources_deep"

    if root == "concepts":
        return "concepts"

    if root == "guides":
        return "guides"

    if root == "programs":
        return "programs"

    if root == "subjects":
        return "subjects"

    if root == "gurgaon-area":
        return "area_cluster"

    if root == "gurugram":
        if len(segments) == 1:
            return "gurugram_hub"
        if segments[1] == "sectors":
            if len(segments) == 2:
                return "sector_hub"
            if len(segments) == 3:
                return "sector_page"
            return "sector_detail"
        if segments[1] == "schools":
            if len(segments) == 2:
                return "school_hub_legacy"
            return "school_service"
        if segments[1] == "societies":
            if len(segments) == 2:
                return "society_hub"
            return "society_service"
        if len(segments) == 2:
            return "gurugram_support"
        return "gurugram_other"

    return "core_support"


def text_or_none(value):
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def main() -> int:
    root = Path(__file__).resolve().parents[1]
    default_source = Path.home() / "Downloads" / "Book1.xlsx"
    source_path = Path(sys.argv[1]).expanduser() if len(sys.argv) > 1 else default_source
    output_path = (
        Path(sys.argv[2]).expanduser()
        if len(sys.argv) > 2
        else root / "apps" / "frontend" / "src" / "data" / "generated" / "page-manifest.json"
    )

    if not source_path.exists():
        raise FileNotFoundError(f"Workbook not found: {source_path}")

    workbook = load_workbook(source_path, read_only=True, data_only=True)
    worksheet = workbook[workbook.sheetnames[0]]
    rows = list(worksheet.iter_rows(values_only=True))
    if not rows:
        raise ValueError("Workbook is empty")

    header = [str(cell) if cell is not None else "" for cell in rows[0]]
    index_by_column = {column: idx for idx, column in enumerate(header)}

    missing_columns = [column for column in REQUIRED_COLUMNS if column not in index_by_column]
    if missing_columns:
        raise ValueError(f"Missing expected columns: {missing_columns}")

    manifest = []
    seen_paths: set[str] = set()

    for row in rows[1:]:
        source_path_value = row[index_by_column["Path"]] if "Path" in index_by_column else None
        source_url_value = row[index_by_column["URL"]] if "URL" in index_by_column else None
        normalized = normalize_path(source_path_value or source_url_value)

        if not normalized or normalized in seen_paths:
            continue

        seen_paths.add(normalized)

        record = {
            "id": row[index_by_column["ID"]],
            "publishingWave": text_or_none(row[index_by_column["Publishing_Wave"]]),
            "url": text_or_none(source_url_value),
            "path": text_or_none(source_path_value),
            "normalizedPath": normalized,
            "routeFamily": route_family(normalized),
            "category": text_or_none(row[index_by_column["Category"]]),
            "tier": text_or_none(row[index_by_column["Tier"]]),
            "priority": text_or_none(row[index_by_column["Priority"]]),
            "board": text_or_none(row[index_by_column["Board"]]),
            "classLabel": text_or_none(row[index_by_column["Class"]]),
            "subject": text_or_none(row[index_by_column["Subject"]]),
            "locationType": text_or_none(row[index_by_column["Location_Type"]]),
            "locationName": text_or_none(row[index_by_column["Location_Name"]]),
            "primaryKeyword": text_or_none(row[index_by_column["Primary_Keyword"]]),
            "searchIntent": text_or_none(row[index_by_column["Search_Intent"]]),
            "wordCountTarget": row[index_by_column["Word_Count_Target"]],
            "indexationAdvice": text_or_none(row[index_by_column["Indexation_Advice"]]),
            "parentHub": text_or_none(row[index_by_column["Parent_Hub"]]),
            "links": [
                text_or_none(row[index_by_column["Link_1"]]),
                text_or_none(row[index_by_column["Link_2"]]),
                text_or_none(row[index_by_column["Link_3"]]),
                text_or_none(row[index_by_column["Link_4"]]),
                text_or_none(row[index_by_column["Link_5"]]),
            ],
            "suggestedSchema": text_or_none(row[index_by_column["Suggested_Schema"]]),
            "contentAngle": text_or_none(row[index_by_column["Content_Angle"]]),
            "uniquenessBlock": text_or_none(row[index_by_column["Uniqueness_Block"]]),
            "sourceUrl": text_or_none(row[index_by_column["Source_URL"]]),
        }
        manifest.append(record)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    with output_path.open("w", encoding="utf-8") as output_file:
        json.dump(manifest, output_file, indent=2, ensure_ascii=False)

    print(f"Wrote {len(manifest)} manifest records to {output_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
